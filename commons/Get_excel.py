# -*- coding: utf-8 -*-
from xlutils.copy import copy
import xlwt
import xlsxwriter
import xlrd
from openpyxl import load_workbook
import datetime
import os
import readConfig as readConfig
import commons.get_file
import openpyxl
proDir = readConfig.proDir
#返回一个dict
def read_casefile(filepach,sheet='Sheet1'):
    testcase = {}
    casefile = xlrd.open_workbook(filepach)
    test_table = casefile.sheet_by_name(sheet)
    row_num = test_table.nrows  # 获取行数
    col_num = test_table.ncols  # 获取列数
    # print(col_num)
    for i in range(row_num):
        if i != 0 and test_table.row_values(i)[2]!='' :
            sbbh = str(test_table.row_values(i)[1])#把第二列的值当key，当第二列的key不会相同时，可用此方法
            testcase[sbbh] = {}
            num=0
            while num<int(col_num):#其他列的内容当做value添加到testcase字典里，格式为列名为key，内容为value
                testcase[sbbh][test_table.row_values(0)[num]] =test_table.row_values(i)[num]
                num+=1
    # print(testcase)
    return testcase
#返回一个list
def read_casefile_2(xls_name, sheet='Sheet1'):
    testcase = []
    xlsxPath = os.path.join(proDir, "testFile", 'case', xls_name)
    casefile = xlrd.open_workbook(xlsxPath)
    test_table = casefile.sheet_by_name(sheet)
    row_num = test_table.nrows  # 获取行数
    col_num = test_table.ncols  # 获取列数
    # print(col_num)
    for i in range(row_num):
        if i != 0:
            a = {}
            num=0
            while num<int(col_num):
                a[test_table.row_values(0)[num]]=test_table.row_values(i)[num]
                num+=1
            testcase.append(a)
    # print(testcase)
    return testcase
# a=read_casefile_2('gzfxCase.xlsx','getgzsjlist')
# print(a)

#向filepach.xlsx里hang行,lie列插入数据datas
def inster_data(filepach,hang,lie,datas):
    casefile = load_workbook(filepach)
    test_table = casefile.active
    test_table.cell(hang,lie,datas)
    casefile.save(filepach)

#添加本地图片超链接
def add_Hyperlinks(filepach,sheet):
    rb = xlrd.open_workbook(filepach)
    wb = copy(rb)
    ws = rb.sheet_by_name(sheet)
    ws1=wb.get_sheet(0)
    row_num = ws.nrows
    for i in range(row_num):
        name =ws.row_values(i)[2]
        link = 'HYPERLINK("./%s";"%s")'%(name,name)
        ws1.write(i, 2, xlwt.Formula(link))
    wb.save(filepach[0:-4]+"xls")
    # wb.save(filepach)


#创建新excel的类
class New_xlsx():

    def __init__(self,path="C:\\Users\\admin\\Desktop\\",name="分析结果"):
        self.path=path
        self.name=name+datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        self.workbook = xlsxwriter.Workbook('%s%s.xlsx'%(self.path,self.name))
        self.worksheet = self.workbook.add_worksheet()

    # 在lie，hang添加data文本数据
    def insetr_datas(self,lie,hang,data):
        ws = self.worksheet
        ws.write(hang,lie, data)

    #在lie，hang添加link超链接，并且以data文本显示
    def insetr_datas_url(self,lie,hang,data,url):
        ws=self.worksheet
        link = '=HYPERLINK("%s","%s")' % (url,data)
        ws.write(hang,lie,link)
    #插入图片，x和y是图片的横纵缩放比，tu_path是图片的本地路径
    def insetr_tp(self,lie,hang,tu_path,x=1,y=1):
        ws = self.worksheet
        ws.insert_image(hang,lie, tu_path, {'x_scale': x, 'y_scale': y})

    #保存Excel
    def save_excel(self):
        self.workbook.close()



